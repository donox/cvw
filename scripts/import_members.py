"""
One-time import of CVW membership data from Excel spreadsheet.

Usage:
    source .venv/bin/activate
    python scripts/import_members.py "/path/to/Year 2026 Membership.xlsx"

Deduplicates against existing DB records by email, or last_name+first_name
if no email is present. Updates existing records; creates new ones.
"""
import sys
from datetime import datetime
from pathlib import Path

# Allow running from project root
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from app.database import SessionLocal
import app.models.program   # noqa — register ProgramComment with SQLAlchemy
import app.models.officer
import app.models.org
import app.models.financial
import app.models.user
from app.models.member import Member


def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(val), fmt).date()
        except ValueError:
            pass
    return None


def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


_seen_emails: set = set()


def find_existing(db, email, last_name, first_name):
    if email:
        m = db.query(Member).filter(Member.email == email).first()
        if m:
            return m
    return (
        db.query(Member)
        .filter(Member.last_name == last_name, Member.first_name == first_name)
        .first()
    )


def claim_email(email):
    """Return email if not yet used in this import run, else None."""
    if not email:
        return None
    if email in _seen_emails:
        print(f"  WARNING: duplicate email '{email}' — importing without email")
        return None
    _seen_emails.add(email)
    return email


def apply_fields(member, fields):
    for k, v in fields.items():
        if v is not None:
            setattr(member, k, v)
        elif getattr(member, k, None) is None:
            setattr(member, k, v)


def import_regular(ws, db):
    created = updated = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        last_name = to_str(row[0])
        if not last_name:
            continue  # totals / blank rows

        first_name  = to_str(row[1])
        address     = to_str(row[2])
        city        = to_str(row[3])
        state       = to_str(row[4])
        zip_code    = to_str(row[5])
        home_phone  = to_str(row[6])
        mobile_phone = to_str(row[7])
        email       = claim_email(to_str(row[8]))
        payment_method = to_str(row[9])
        date_paid   = parse_date(row[10])
        dues_raw    = row[11]
        donation_raw = row[12]

        dues_amount = float(dues_raw) if isinstance(dues_raw, (int, float)) else None
        donation    = float(donation_raw) if isinstance(donation_raw, (int, float)) else None
        dues_paid   = bool(dues_amount and dues_amount > 0)

        fields = dict(
            last_name=last_name,
            first_name=first_name,
            address=address,
            city=city,
            state=state,
            zip_code=zip_code,
            home_phone=home_phone,
            mobile_phone=mobile_phone,
            email=email,
            payment_method=payment_method,
            dues_paid_date=date_paid,
            dues_amount=dues_amount,
            donation=donation,
            dues_paid=dues_paid,
            membership_type="Individual",
            status="Active",
        )

        existing = find_existing(db, email, last_name, first_name)
        if existing:
            apply_fields(existing, fields)
            updated += 1
        else:
            db.add(Member(**{k: v for k, v in fields.items() if v is not None}))
            created += 1

    return created, updated


def import_affiliated(ws, db):
    created = updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        last_name = to_str(row[0])
        if not last_name:
            continue

        first_name   = to_str(row[1])
        address      = to_str(row[2])
        city         = to_str(row[3])
        state        = to_str(row[4])
        zip_code     = to_str(row[5])
        home_phone   = to_str(row[6])
        mobile_phone = to_str(row[7])
        email        = claim_email(to_str(row[8]))
        payment_method = to_str(row[9])
        date_paid    = parse_date(row[10])
        dues_raw     = row[11]

        dues_amount = float(dues_raw) if isinstance(dues_raw, (int, float)) else None
        dues_paid   = bool(dues_amount and dues_amount > 0)

        fields = dict(
            last_name=last_name,
            first_name=first_name,
            address=address,
            city=city,
            state=state,
            zip_code=zip_code,
            home_phone=home_phone,
            mobile_phone=mobile_phone,
            email=email,
            payment_method=payment_method,
            dues_paid_date=date_paid,
            dues_amount=dues_amount,
            dues_paid=dues_paid,
            membership_type="Secondary/Affiliate",
            status="Active",
        )

        existing = find_existing(db, email, last_name, first_name)
        if existing:
            apply_fields(existing, fields)
            updated += 1
        else:
            db.add(Member(**{k: v for k, v in fields.items() if v is not None}))
            created += 1

    return created, updated


def import_honorary(ws, db):
    created = updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        last_name = to_str(row[1])
        if not last_name:
            continue

        first_name   = to_str(row[2])
        address      = to_str(row[3])
        city         = to_str(row[4])
        state        = to_str(row[5])
        zip_code     = to_str(row[6])
        home_phone   = to_str(row[7])
        mobile_phone = to_str(row[8])
        email        = claim_email(to_str(row[9]))

        fields = dict(
            last_name=last_name,
            first_name=first_name,
            address=address,
            city=city,
            state=state,
            zip_code=zip_code,
            home_phone=home_phone,
            mobile_phone=mobile_phone,
            email=email,
            membership_type="Life",
            status="Active",
            dues_paid=False,
        )

        existing = find_existing(db, email, last_name, first_name)
        if existing:
            apply_fields(existing, fields)
            updated += 1
        else:
            db.add(Member(**{k: v for k, v in fields.items() if v is not None}))
            created += 1

    return created, updated


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "/home/don/Documents/CVW/Year 2026 Membership.xlsx"
    wb = openpyxl.load_workbook(path)

    db = SessionLocal()
    try:
        # Seed seen-emails with what's already in DB so claim_email works correctly
        existing_emails = {e for (e,) in db.query(Member.email).filter(Member.email.isnot(None)).all()}
        _seen_emails.update(existing_emails)

        r_c, r_u = import_regular(wb["Regular Members"], db)
        a_c, a_u = import_affiliated(wb["Affiliated Members"], db)
        h_c, h_u = import_honorary(wb["Honorary Members"], db)
        db.commit()

        total_c = r_c + a_c + h_c
        total_u = r_u + a_u + h_u
        print(f"Regular Members:   {r_c} created, {r_u} updated")
        print(f"Affiliated:        {a_c} created, {a_u} updated")
        print(f"Honorary:          {h_c} created, {h_u} updated")
        print(f"Total:             {total_c} created, {total_u} updated")
    except Exception as e:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
